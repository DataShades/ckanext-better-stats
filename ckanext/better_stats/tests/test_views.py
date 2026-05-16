from __future__ import annotations

import csv
import io
import json
from typing import Any
from unittest import mock

import pytest
from openpyxl import load_workbook

from ckanext.better_stats import const
from ckanext.better_stats.metrics.base import MetricBase, MetricRegistry
from ckanext.better_stats.model import MetricConfig, UserFavorite
from ckanext.better_stats.views.dashboard import MetricExporter


class PublicMetric(MetricBase):
    supported_visualizations = [const.VisualizationType.TABLE]
    default_visualization = const.VisualizationType.TABLE

    def __init__(self) -> None:
        super().__init__(name="public_metric", title="Public", access_level="public")

    def get_data(self) -> int:
        return 5

    def get_table_data(self) -> dict[str, Any]:
        return {"headers": ["A", "B"], "rows": [["x", 1], ["y", 2]]}


class AdminMetric(MetricBase):
    supported_visualizations = [const.VisualizationType.CARD]
    default_visualization = const.VisualizationType.CARD

    def __init__(self) -> None:
        super().__init__(name="admin_metric", title="Admin", access_level="admin")

    def get_data(self) -> int:
        return 10

@pytest.fixture
def public_metric_only(fresh_registry: Any) -> Any:
    MetricRegistry._loaded = True
    MetricRegistry.register("public_metric", PublicMetric)
    return


@pytest.fixture
def admin_metric_only(fresh_registry: Any) -> Any:
    MetricRegistry._loaded = True
    MetricRegistry.register("admin_metric", AdminMetric)
    return


@pytest.mark.usefixtures("with_plugins")
class TestMetricExporter:
    def _metric(self) -> PublicMetric:
        return PublicMetric()

    def test_csv_export(self, app: Any) -> None:
        with app.flask_app.test_request_context("/x"):
            response = MetricExporter(
                self._metric(), "out", const.ExportFormat.CSV.value
            ).export_metric()
        assert response.headers["Content-Type"] == "text/csv"
        assert "out.csv" in response.headers["Content-Disposition"]

        rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
        assert rows[0] == ["A", "B"]
        assert rows[1] == ["x", "1"]

    def test_json_export(self, app: Any) -> None:
        with app.flask_app.test_request_context("/x"):
            response = MetricExporter(
                self._metric(), "out", const.ExportFormat.JSON.value
            ).export_metric()
        assert response.headers["Content-Type"] == "application/json"
        envelope = json.loads(response.get_data(as_text=True))
        assert envelope["metric"] == "public_metric"
        assert envelope["title"] == "Public"
        assert envelope["data"]["headers"] == ["A", "B"]

    def test_xlsx_export(self, app: Any) -> None:
        with app.flask_app.test_request_context("/x"):
            response = MetricExporter(
                self._metric(), "out", const.ExportFormat.XLSX.value
            ).export_metric()
        assert response.headers["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument"
        )
        wb = load_workbook(io.BytesIO(response.get_data()))
        ws = wb.active
        assert ws is not None
        assert [c.value for c in ws[1]] == ["A", "B"]
        assert ws.cell(row=2, column=1).value == "x"

    def test_xlsx_export_handles_dict_cells(self, app: Any) -> None:
        class TableWithLinks(PublicMetric):
            def get_table_data(self) -> dict[str, Any]:
                return {
                    "headers": ["Col"],
                    "rows": [[{"text": "Hello", "url": "/x"}]],
                }

        with app.flask_app.test_request_context("/x"):
            response = MetricExporter(
                TableWithLinks(), "out", const.ExportFormat.XLSX.value
            ).export_metric()
        wb = load_workbook(io.BytesIO(response.get_data()))
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=2, column=1).value == "Hello"

    def test_unsupported_format_returns_400(self, app: Any) -> None:
        with app.flask_app.test_request_context("/x"):
            response = MetricExporter(self._metric(), "out", "pdf").export_metric()
        assert response.status_code == 400


@pytest.mark.usefixtures("with_plugins", "clean_db")
class TestDashboardEndpoints:
    def test_dashboard_renders(self, app: Any) -> None:
        response = app.get("/better_stats/dashboard")
        assert response.status_code == 200

    def test_get_metric_data_unknown(self, app: Any) -> None:
        response = app.get("/better_stats/metric/missing")
        assert response.status_code == 404

    def test_get_metric_data_success(self, app: Any, public_metric_only: Any) -> None:
        response = app.get("/better_stats/metric/public_metric")
        assert response.status_code == 200
        body = json.loads(response.body)
        assert body["name"] == "public_metric"
        assert body["data"]["headers"] == ["A", "B"]

    def test_get_metric_data_invalid_viz_falls_back(
        self, app: Any, public_metric_only: Any
    ) -> None:
        response = app.get("/better_stats/metric/public_metric?type=bogus")
        body = json.loads(response.body)
        assert body["type"] == const.VisualizationType.TABLE.value

    def test_get_metric_data_unsupported_viz_falls_back(
        self, app: Any, public_metric_only: Any
    ) -> None:
        response = app.get("/better_stats/metric/public_metric?type=chart")
        body = json.loads(response.body)
        assert body["type"] == const.VisualizationType.TABLE.value

    def test_get_metric_data_refresh(self, app: Any, public_metric_only: Any) -> None:
        response = app.get("/better_stats/metric/public_metric?refresh=true")
        assert response.status_code == 200

    def test_get_metrics_batch_requires_names(self, app: Any) -> None:
        response = app.get("/better_stats/metrics", status=400)
        assert response.status_code == 400

    def test_get_metrics_batch_returns_results_and_errors(
        self, app: Any, public_metric_only: Any
    ) -> None:
        response = app.get("/better_stats/metrics?names=public_metric,missing")
        body = json.loads(response.body)
        assert "public_metric" in body["metrics"]
        assert "missing" in body["errors"]

    def test_export_metric_unknown(self, app: Any) -> None:
        response = app.get("/better_stats/export/missing", status=404)
        assert response.status_code == 404

    def test_export_metric_unsupported_format(
        self, app: Any, public_metric_only: Any
    ) -> None:
        response = app.get("/better_stats/export/public_metric?format=pdf", status=400)
        assert response.status_code == 400

    def test_export_metric_csv(self, app: Any, public_metric_only: Any) -> None:
        response = app.get("/better_stats/export/public_metric?format=csv")
        assert response.status_code == 200
        assert response.headers["Content-Type"] == "text/csv"

    def test_export_metric_json(self, app: Any, public_metric_only: Any) -> None:
        response = app.get("/better_stats/export/public_metric?format=json")
        assert response.status_code == 200
        assert response.headers["Content-Type"] == "application/json"

    def test_export_metric_inaccessible(self, app: Any, admin_metric_only: Any) -> None:
        response = app.get("/better_stats/export/admin_metric", status=403)
        assert response.status_code == 403

    def test_embed_metric_unknown_returns_404(self, app: Any) -> None:
        response = app.get("/better_stats/embed/missing", status=404)
        assert response.status_code == 404

    def test_embed_metric_renders(self, app: Any, public_metric_only: Any) -> None:
        response = app.get("/better_stats/embed/public_metric")
        assert response.status_code == 200
        assert response.headers["X-Frame-Options"] == "ALLOWALL"

    def test_toggle_favorite_anonymous(self, app: Any, public_metric_only: Any) -> None:
        response = app.post("/better_stats/favorites/toggle/public_metric", status=401)
        assert response.status_code == 401

    def test_toggle_favorite_unknown_metric(
        self, app: Any, sysadmin: dict[str, Any]
    ) -> None:
        response = app.post(
            "/better_stats/favorites/toggle/missing",
            headers={"Authorization": sysadmin["token"]},
            status=404,
        )
        assert response.status_code == 404

    def test_toggle_favorite_round_trip(
        self,
        app: Any,
        sysadmin: dict[str, Any],
        public_metric_only: Any,
    ) -> None:
        headers = {"Authorization": sysadmin["token"]}

        response = app.post(
            "/better_stats/favorites/toggle/public_metric", headers=headers
        )
        assert response.status_code == 200
        body = json.loads(response.body)
        assert body["is_favorite"] is True
        assert UserFavorite.get(sysadmin["id"], "public_metric") is not None

        response = app.post(
            "/better_stats/favorites/toggle/public_metric", headers=headers
        )
        body = json.loads(response.body)
        assert body["is_favorite"] is False
        assert UserFavorite.get(sysadmin["id"], "public_metric") is None


@pytest.mark.usefixtures("with_plugins")
class TestSettingsEndpoints:
    def test_settings_anonymous_forbidden(self, app: Any) -> None:
        response = app.get("/better_stats/settings", status=403)
        assert response.status_code == 403

    def test_settings_user_forbidden(self, app: Any, user: dict[str, Any]) -> None:
        response = app.get(
            "/better_stats/settings",
            headers={"Authorization": user["token"]},
            status=403,
        )
        assert response.status_code == 403

    def test_settings_sysadmin_allowed(
        self, app: Any, sysadmin: dict[str, Any]
    ) -> None:
        response = app.get(
            "/better_stats/settings",
            headers={"Authorization": sysadmin["token"]},
        )
        assert response.status_code == 200

    def test_clear_metric_cache_unknown(
        self, app: Any, sysadmin: dict[str, Any]
    ) -> None:
        response = app.post(
            "/better_stats/settings/cache/clear/missing",
            headers={"Authorization": sysadmin["token"]},
            status=404,
        )
        assert response.status_code == 404

    def test_clear_metric_cache(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        with mock.patch("ckanext.better_stats.metrics.base.cache"):
            response = app.post(
                "/better_stats/settings/cache/clear/public_metric",
                headers={"Authorization": sysadmin["token"]},
            )
        assert response.status_code == 200
        assert json.loads(response.body)["cleared"] == "public_metric"

    def test_clear_all_caches(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        with mock.patch("ckanext.better_stats.metrics.base.cache"):
            response = app.post(
                "/better_stats/settings/cache/clear",
                headers={"Authorization": sysadmin["token"]},
            )
        body = json.loads(response.body)
        assert body["cleared"] == ["public_metric"]

    def test_reset_all_configs(self, app: Any, sysadmin: dict[str, Any]) -> None:
        MetricConfig.upsert("any_metric", enabled=False)
        response = app.post(
            "/better_stats/settings/reset",
            headers={"Authorization": sysadmin["token"]},
        )
        assert json.loads(response.body)["reset"] is True
        assert MetricConfig.for_metric("any_metric") is None

    def test_update_metric_config_validation_error(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        response = app.post(
            "/better_stats/settings/metric/public_metric",
            data=json.dumps({"col_span": 99}),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
            status=400,
        )
        assert response.status_code == 400

    def test_update_metric_config_not_found(
        self, app: Any, sysadmin: dict[str, Any]
    ) -> None:
        response = app.post(
            "/better_stats/settings/metric/missing",
            data=json.dumps({"col_span": 2}),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
            status=404,
        )
        assert response.status_code == 404

    def test_update_metric_config_success(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        response = app.post(
            "/better_stats/settings/metric/public_metric",
            data=json.dumps({"col_span": 4, "enabled": False}),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
        )
        body = json.loads(response.body)
        assert body["col_span"] == 4
        assert body["enabled"] is False

    def test_batch_update_order_invalid_payload(
        self, app: Any, sysadmin: dict[str, Any]
    ) -> None:
        response = app.post(
            "/better_stats/settings/batch-order",
            data=json.dumps({"not": "a list"}),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
            status=400,
        )
        assert response.status_code == 400

    def test_batch_update_order_unknown_metric(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        response = app.post(
            "/better_stats/settings/batch-order",
            data=json.dumps([{"metric_name": "missing", "order": 1}]),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
            status=400,
        )
        assert response.status_code == 400

    def test_batch_update_order_invalid_order(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        response = app.post(
            "/better_stats/settings/batch-order",
            data=json.dumps([{"metric_name": "public_metric", "order": "bad"}]),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
            status=400,
        )
        assert response.status_code == 400

    def test_batch_update_order_success(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        response = app.post(
            "/better_stats/settings/batch-order",
            data=json.dumps([{"metric_name": "public_metric", "order": 999}]),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
        )
        assert response.status_code == 200
        cfg = MetricConfig.for_metric("public_metric")
        assert cfg is not None
        assert cfg.order == 999

    def test_batch_update_order_updates_existing(
        self, app: Any, sysadmin: dict[str, Any], public_metric_only: Any
    ) -> None:
        MetricConfig.upsert("public_metric", order=1)
        response = app.post(
            "/better_stats/settings/batch-order",
            data=json.dumps([{"metric_name": "public_metric", "order": 50}]),
            content_type="application/json",
            headers={"Authorization": sysadmin["token"]},
        )
        assert response.status_code == 200
        cfg = MetricConfig.for_metric("public_metric")
        assert cfg is not None
        assert cfg.order == 50
